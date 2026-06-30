"""Relatório de Fechamento: agrega tentativas por data, aplica config/exceções e RF ENTRADA."""

from __future__ import annotations

from collections import defaultdict
from datetime import date
from decimal import Decimal

from django.db.models import Count

from pages.financeiro.models import RegistroFinanceiro, RegistroFinanceiroTipo
from pages.logistica_config.models import ConfiguracaoLogistica
from pages.logistica_config.services.excecoes_resolucao import (
    carregar_contexto_excecoes_intervalo,
    resolver_reservados_para_data,
)
from pages.pedidos.models import TentativaEntrega

PERIODO_MAXIMO_DIAS = 90

_DIAS_SEMANA_PT = (
    "segunda-feira",
    "terça-feira",
    "quarta-feira",
    "quinta-feira",
    "sexta-feira",
    "sábado",
    "domingo",
)


def formatar_data_coluna_fechamento(d: date) -> str:
    """Ex.: terça-feira, 12/05/2026 (weekday alinhado a `date.weekday()`, segunda=0)."""
    return f"{_DIAS_SEMANA_PT[d.weekday()]}, {d.strftime('%d/%m/%Y')}"


def formatar_inteiro_pt_br(n: int) -> str:
    """Agrupa milhares com ponto (ex.: 11111 -> '11.111')."""
    neg = n < 0
    x = abs(int(n))
    s = str(x)
    partes = []
    while len(s) > 3:
        partes.append(s[-3:])
        s = s[:-3]
    partes.append(s)
    corpo = ".".join(reversed(partes))
    return ("-" if neg else "") + corpo


def formatar_decimal_pt_br(valor: Decimal, casas_decimais: int = 2) -> str:
    """Formato monetário pt-BR: milhar com ponto, decimais com vírgula (ex.: 11.111,11)."""
    q = valor.quantize(Decimal(10) ** -casas_decimais)
    neg = q < 0
    q = abs(q)
    texto = format(q, f".{casas_decimais}f")
    parte_int, parte_frac = texto.split(".")
    # parte_int sem sinal; evita duplo '-' em negativos
    ip_num = int(parte_int)
    ip_fmt = formatar_inteiro_pt_br(ip_num)
    return ("-" if neg else "") + ip_fmt + "," + parte_frac


def parse_inteiro_pt_br(texto) -> int | None:
    """Converte inteiro formatado pt-BR (ex.: '11.111') para int."""
    if texto is None or texto == "":
        return None
    s = str(texto).strip().replace(".", "")
    if not s or s == "-":
        return None
    return int(s)


def parse_decimal_pt_br(texto) -> Decimal | None:
    """Converte decimal formatado pt-BR (ex.: '11.111,11') para Decimal."""
    if texto is None or texto == "":
        return None
    s = str(texto).strip()
    if not s:
        return None
    neg = s.startswith("-")
    s = s.lstrip("-").replace(".", "").replace(",", ".")
    if not s:
        return None
    d = Decimal(s)
    return -d if neg else d


def _item_expresso_linha_payload(item: dict) -> dict:
    """Monta o dict de exibição a partir do item em rf_por_data (só obs + valor formatado)."""
    raw_valor = item.get("valor", "0")
    return {
        "observacao": (item.get("observacao") or ""),
        "valor": formatar_decimal_pt_br(Decimal(str(raw_valor))),
    }


def validar_periodo(data_ini: date, data_fim: date) -> str | None:
    if data_ini > data_fim:
        return "A data inicial não pode ser maior que a data final."
    if (data_fim - data_ini).days > PERIODO_MAXIMO_DIAS:
        return f"O período máximo é de {PERIODO_MAXIMO_DIAS} dias."
    return None


def montar_relatorio_fechamento(filial, data_ini: date, data_fim: date) -> tuple[dict | None, str | None]:
    cfg = ConfiguracaoLogistica.objects.filter(filial=filial).first()
    if not cfg:
        return None, (
            "Não existe configuração de logística para a filial ativa. "
            "Cadastre-a em Logística > Cadastro > Configuração de Logística antes de emitir o relatório."
        )

    agg_rows = (
        TentativaEntrega.objects.filter(
            pedido__filial=filial,
            pedido__expresso=False,
            data_tentativa__range=(data_ini, data_fim),
        )
        .values("data_tentativa")
        .annotate(qtd_pedidos=Count("pedido_id", distinct=True))
        .order_by("data_tentativa")
    )

    exce_individual, periodos_excecao = carregar_contexto_excecoes_intervalo(cfg, data_ini, data_fim)

    rf_por_data: dict[date, list[dict]] = defaultdict(list)
    for r in RegistroFinanceiro.objects.filter(
        filial=filial,
        tipo=RegistroFinanceiroTipo.ENTRADA,
        data_emissao__range=(data_ini, data_fim),
    ).values("id", "data_emissao", "observacao", "valor"):
        vd = r["valor"]
        if not isinstance(vd, Decimal):
            vd = Decimal(str(vd))
        rf_por_data[r["data_emissao"]].append(
            {
                "observacao": (r["observacao"] or ""),
                "valor": str(vd),
            }
        )

    v_unit_l = cfg.valor_unitario_ligeiro
    v_unit_p = cfg.valor_unitario_pesado

    tot_qtd = 0
    tot_l_q = 0
    tot_l_v = Decimal("0")
    tot_p_q = 0
    tot_p_v = Decimal("0")
    tot_res = 0
    tot_expresso_v = Decimal("0")
    tot_exc = 0

    linhas = []
    for row in agg_rows:
        d = row["data_tentativa"]
        qtd = row["qtd_pedidos"]
        qty_l, qty_p = resolver_reservados_para_data(d, cfg, exce_individual, periodos_excecao)

        valor_l = (Decimal(qty_l) * v_unit_l).quantize(Decimal("0.01"))
        valor_p = (Decimal(qty_p) * v_unit_p).quantize(Decimal("0.01"))

        ped_reservados = int(cfg.pedidos_ligeiro) * int(qty_l) + int(cfg.pedidos_pesado) * int(qty_p)
        ped_excedentes = max(0, int(qtd) - ped_reservados)

        expresso_raw = rf_por_data.get(d, [])
        for ex_item in expresso_raw:
            tot_expresso_v += Decimal(str(ex_item.get("valor", "0")))

        tot_qtd += int(qtd)
        tot_l_q += int(qty_l)
        tot_l_v += valor_l
        tot_p_q += int(qty_p)
        tot_p_v += valor_p
        tot_res += ped_reservados
        tot_exc += ped_excedentes

        expresso_fmt = [_item_expresso_linha_payload(x) for x in expresso_raw]

        linhas.append(
            {
                "data": formatar_data_coluna_fechamento(d),
                "qtd_pedidos": formatar_inteiro_pt_br(int(qtd)),
                "ligeiro_quantidade": formatar_inteiro_pt_br(int(qty_l)),
                "ligeiro_valor": formatar_decimal_pt_br(valor_l),
                "pesado_quantidade": formatar_inteiro_pt_br(int(qty_p)),
                "pesado_valor": formatar_decimal_pt_br(valor_p),
                "pedidos_reservados": formatar_inteiro_pt_br(ped_reservados),
                "pedidos_excedentes": formatar_inteiro_pt_br(ped_excedentes),
                "expresso": expresso_fmt,
            }
        )

    tot_exc_valor = (Decimal(tot_exc) * cfg.valor_excedente).quantize(Decimal("0.01"))

    valor_total_consolidado = (tot_l_v + tot_p_v + tot_exc_valor + tot_expresso_v).quantize(Decimal("0.01"))
    periodo_texto = f"{data_ini.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"

    totais = {
        "qtd_pedidos": formatar_inteiro_pt_br(tot_qtd),
        "ligeiro_quantidade": formatar_inteiro_pt_br(tot_l_q),
        "ligeiro_valor": formatar_decimal_pt_br(tot_l_v),
        "pesado_quantidade": formatar_inteiro_pt_br(tot_p_q),
        "pesado_valor": formatar_decimal_pt_br(tot_p_v),
        "pedidos_reservados": formatar_inteiro_pt_br(tot_res),
        "pedidos_excedentes": formatar_inteiro_pt_br(tot_exc),
        "pedidos_excedentes_valor": formatar_decimal_pt_br(tot_exc_valor),
        "expresso_valor": formatar_decimal_pt_br(tot_expresso_v),
    }

    return {
        "linhas": linhas,
        "totais": totais,
        "periodo_texto": periodo_texto,
        "valor_total_consolidado": formatar_decimal_pt_br(valor_total_consolidado),
        "periodo_maximo_dias": PERIODO_MAXIMO_DIAS,
    }, None


_RF_SEP = " | "
_RF_SEP_OBS_VALOR = " — "
_FMT_DECIMAL_XLSX = "#,##0.00"


def _celula_int_xlsx(valor_fmt):
    n = parse_inteiro_pt_br(valor_fmt)
    return n if n is not None else None


def _celula_decimal_xlsx(valor_fmt):
    d = parse_decimal_pt_br(valor_fmt)
    return float(d) if d is not None else None


def _formatar_lancamento_expresso_export(item: dict) -> str:
    obs = (item.get("observacao") or "").strip()
    val = item.get("valor")
    val_s = str(val) if val not in (None, "") else ""
    if obs:
        return f"{obs}{_RF_SEP_OBS_VALOR}{val_s}"
    return val_s


def _juntar_lista_expresso_export(lista) -> str:
    if not lista:
        return ""
    partes = [_formatar_lancamento_expresso_export(x) for x in lista]
    return _RF_SEP.join(p for p in partes if p)


def gerar_xlsx_relatorio_fechamento(payload: dict) -> bytes:
    """Gera workbook .xlsx com colunas numéricas para quantidades e valores."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Fechamento"

    periodo = payload.get("periodo_texto") or ""
    valor_total = payload.get("valor_total_consolidado") or ""
    if periodo:
        ws.append([f"Período: {periodo} — {valor_total}"])
        ws["A1"].font = Font(bold=True)
        ws.append([])

    cabecalhos = [
        "Data",
        "Pedidos",
        "Ligeiro Qtd",
        "Ligeiro Valor",
        "Pesado Qtd",
        "Pesado Valor",
        "Reservados",
        "Excedentes",
        "Excedentes Valor",
        "Expresso (Obs. / Valor)",
    ]
    ws.append(cabecalhos)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = Font(bold=True)

    colunas_decimais = {4, 6, 9, 10}

    for linha in payload.get("linhas") or []:
        row_idx = ws.max_row + 1
        ws.append([
            linha.get("data") or "",
            _celula_int_xlsx(linha.get("qtd_pedidos")),
            _celula_int_xlsx(linha.get("ligeiro_quantidade")),
            _celula_decimal_xlsx(linha.get("ligeiro_valor")),
            _celula_int_xlsx(linha.get("pesado_quantidade")),
            _celula_decimal_xlsx(linha.get("pesado_valor")),
            _celula_int_xlsx(linha.get("pedidos_reservados")),
            _celula_int_xlsx(linha.get("pedidos_excedentes")),
            None,
            _juntar_lista_expresso_export(linha.get("expresso")),
        ])
        for col in colunas_decimais:
            cell = ws.cell(row=row_idx, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = _FMT_DECIMAL_XLSX

    totais = payload.get("totais")
    if totais:
        row_idx = ws.max_row + 1
        ws.append([
            "Totais",
            _celula_int_xlsx(totais.get("qtd_pedidos")),
            _celula_int_xlsx(totais.get("ligeiro_quantidade")),
            _celula_decimal_xlsx(totais.get("ligeiro_valor")),
            _celula_int_xlsx(totais.get("pesado_quantidade")),
            _celula_decimal_xlsx(totais.get("pesado_valor")),
            _celula_int_xlsx(totais.get("pedidos_reservados")),
            _celula_int_xlsx(totais.get("pedidos_excedentes")),
            _celula_decimal_xlsx(totais.get("pedidos_excedentes_valor")),
            _celula_decimal_xlsx(totais.get("expresso_valor")),
        ])
        for cell in ws[row_idx]:
            cell.font = Font(bold=True)
        for col in colunas_decimais:
            cell = ws.cell(row=row_idx, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = _FMT_DECIMAL_XLSX

    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 48)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
